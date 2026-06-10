from typing import Dict, List, Optional, Tuple
from pathlib import Path


class EquivalenceDb:
    def __init__(self):
        self._map: Dict[str, str] = {}
        self._reverse_map: Dict[str, str] = {}

    @classmethod
    def from_excel(cls, path: str) -> "EquivalenceDb":
        import openpyxl
        db = cls()
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        header = [str(c).strip().lower() if c else '' for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        try:
            can_idx = header.index('nombre_can')
            var_idx = header.index('variable_interna')
        except ValueError as e:
            raise ValueError(
                "El Excel de equivalencias debe tener las columnas "
                "'nombre_can' y 'variable_interna'"
            ) from e
        for row in ws.iter_rows(min_row=2, values_only=True):
            can = row[can_idx]
            var = row[var_idx]
            if can is None or var is None:
                continue
            can = str(can).strip()
            var = str(var).strip()
            if can and var:
                db._map[can] = var
                db._reverse_map[var] = can
        return db

    def add(self, nombre_can: str, variable_interna: str):
        self._map[nombre_can] = variable_interna
        self._reverse_map[variable_interna] = nombre_can

    def get(self, nombre_can: str) -> Optional[str]:
        return self._map.get(nombre_can)

    def get_can_name(self, variable_interna: str) -> Optional[str]:
        return self._reverse_map.get(variable_interna)

    def __contains__(self, nombre_can: str) -> bool:
        return nombre_can in self._map

    def __len__(self) -> int:
        return len(self._map)

    def all_can_names(self) -> List[str]:
        return list(self._map.keys())

    def all_internal_vars(self) -> List[str]:
        return list(self._map.values())
