from pathlib import Path
from typing import Dict, Optional
from dataclasses import dataclass
from openpyxl import load_workbook


@dataclass
class CalibrationData:
    """Calibration scalar values relevant to torque derating analysis."""
    filepath: str = ''
    cxx_max_norm_tq: float = 1.0
    cxx_min_norm_tq: float = -1.0
    cxx_max_norm_tq_emot2: float = 1.0
    cxx_min_norm_tq_emot2: float = -1.0
    raw: Dict[str, float] = None

    def __post_init__(self):
        if self.raw is None:
            self.raw = {}

    def get(self, label: str, default: Optional[float] = None) -> Optional[float]:
        return self.raw.get(label, default)


def read_calibrations(filepath: str) -> CalibrationData:
    """Read calibration scalar values from the inverter calibration Excel.

    Expected format (Overview sheet):
        Row 1: Document header
        Row 2: Type | Label | Units | Owner | Value
        Row 3+: Data rows

    Returns a CalibrationData with torque-relevant calibratables
    and a raw dict of all scalar values found.
    """
    wb = load_workbook(filepath, data_only=True)
    ws_name = 'Overview' if 'Overview' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[ws_name]

    cxx_max_norm_tq = 1.0
    cxx_min_norm_tq = -1.0
    cxx_max_norm_tq_emot2 = 1.0
    cxx_min_norm_tq_emot2 = -1.0
    raw: Dict[str, float] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 5:
            continue
        label = row[1]
        value = row[4]
        if label is None or not isinstance(label, str):
            continue
        if value is None or isinstance(value, str):
            continue
        try:
            val = float(value)
        except (ValueError, TypeError):
            continue
        raw[label] = val
        if label == 'Cxx_max_norm_tq':
            cxx_max_norm_tq = val
        elif label == 'Cxx_min_norm_tq':
            cxx_min_norm_tq = val
        elif label == 'Cxx_max_norm_tq_emot2':
            cxx_max_norm_tq_emot2 = val
        elif label == 'Cxx_min_norm_tq_emot2':
            cxx_min_norm_tq_emot2 = val

    return CalibrationData(
        filepath=filepath,
        cxx_max_norm_tq=cxx_max_norm_tq,
        cxx_min_norm_tq=cxx_min_norm_tq,
        cxx_max_norm_tq_emot2=cxx_max_norm_tq_emot2,
        cxx_min_norm_tq_emot2=cxx_min_norm_tq_emot2,
        raw=raw,
    )
