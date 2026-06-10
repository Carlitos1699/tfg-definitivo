from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, field
import numpy as np
from openpyxl import load_workbook


@dataclass
class TorqueTable:
    temperatures: List[float] = field(default_factory=list)
    voltages: Dict[float, np.ndarray] = field(default_factory=dict)
    speeds: Dict[float, np.ndarray] = field(default_factory=dict)
    cmax: Dict[float, np.ndarray] = field(default_factory=dict)
    cmin: Dict[float, np.ndarray] = field(default_factory=dict)


class TorquePerfDatabase:
    def __init__(self):
        self._tables: Dict[str, TorqueTable] = {}

    @staticmethod
    def from_excel(filepath: str) -> 'TorquePerfDatabase':
        db = TorquePerfDatabase()
        wb = load_workbook(filepath, data_only=True)
        for sheet_name in ('ME', 'HSG'):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            table = TorqueTable()
            i = 0
            while i < len(rows):
                row = rows[i]
                temp_str = row[1]
                if temp_str is not None and isinstance(temp_str, str) and 'Temp' in temp_str:
                    header = rows[i + 1]
                    is_torque = (header[1] and 'Cmax' in str(header[1]))
                    if not is_torque:
                        i += 1
                        continue
                    temp_val = _parse_temp(temp_str)
                    if temp_val is not None:
                        vrow = rows[i + 2]
                        cmax_voltages, cmax_speeds, cmax_vals = _parse_block(rows, i + 3, vrow, cmax=True)
                        cmin_voltages, cmin_speeds, cmin_vals = _parse_block(rows, i + 3, vrow, cmax=False)
                        if cmax_voltages is not None:
                            table.temperatures.append(temp_val)
                            table.voltages[temp_val] = cmax_voltages
                            table.speeds[temp_val] = cmax_speeds
                            table.cmax[temp_val] = cmax_vals
                            table.cmin[temp_val] = cmin_vals
                i += 1
            db._tables[sheet_name] = table
        return db

    def get_temperatures(self, machine: str) -> List[float]:
        tbl = self._tables.get(machine)
        return list(tbl.temperatures) if tbl else []

    def interpolate(self, machine: str, voltage: float, speed: float,
                    temperature: float = 40.0) -> Tuple[Optional[float], Optional[float]]:
        tbl = self._tables.get(machine)
        if not tbl or not tbl.temperatures:
            return None, None
        temp = _nearest_temp(tbl.temperatures, temperature)
        v_arr = tbl.voltages[temp]
        s_arr = tbl.speeds[temp]
        cmax_arr = tbl.cmax[temp]
        cmin_arr = tbl.cmin[temp]
        cmax = _bilinear(v_arr, s_arr, cmax_arr, voltage, speed)
        cmin = _bilinear(v_arr, s_arr, cmin_arr, voltage, speed)
        return cmin, cmax


def _parse_temp(s: str) -> Optional[float]:
    try:
        for ch in ('Temp. essai = ', 'Temp. essai =', '°C', '�C', '\u00b0C'):
            s = s.replace(ch, '')
        s = s.strip()
        return float(s)
    except (ValueError, AttributeError):
        return None


def _parse_block(rows: list, data_start: int, vrow: tuple, cmax: bool):
    if cmax:
        spd_col = 1
        vol_start = 2
    else:
        spd_col = 11
        vol_start = 12

    voltages = []
    for j in range(vol_start, 19):
        v = vrow[j]
        if v is not None:
            try:
                voltages.append(float(v))
            except (ValueError, TypeError):
                pass
        # If we've found voltages and hit a gap, stop
        elif len(voltages) > 0:
            break
    if not voltages:
        return None, None, None

    n_volts = len(voltages)
    speeds = []
    data = []

    for row in rows[data_start:]:
        speed = row[spd_col]
        if speed is None:
            break
        try:
            speed = float(speed)
        except (ValueError, TypeError):
            continue
        vals = []
        for j in range(vol_start, vol_start + n_volts):
            v = row[j]
            if v is not None:
                try:
                    vals.append(float(v))
                except (ValueError, TypeError):
                    vals.append(np.nan)
            else:
                vals.append(np.nan)
        if len(vals) == n_volts and not all(np.isnan(v) for v in vals):
            speeds.append(speed)
            data.append(vals)

    v_arr = np.array(voltages, dtype=float)
    s_arr = np.array(speeds, dtype=float)
    d_arr = np.array(data, dtype=float)
    return v_arr, s_arr, d_arr


def _nearest_temp(temps: List[float], target: float) -> float:
    return min(temps, key=lambda t: abs(t - target))


def _bilinear(x_vals: np.ndarray, y_vals: np.ndarray, z: np.ndarray,
              x: float, y: float) -> Optional[float]:
    if len(x_vals) < 2 or len(y_vals) < 2:
        return None
    if x < x_vals[0] or x > x_vals[-1] or y < y_vals[0] or y > y_vals[-1]:
        return None
    xi = np.searchsorted(x_vals, x) - 1
    xi = max(0, min(xi, len(x_vals) - 2))
    yi = np.searchsorted(y_vals, y) - 1
    yi = max(0, min(yi, len(y_vals) - 2))
    x0, x1 = x_vals[xi], x_vals[xi + 1]
    y0, y1 = y_vals[yi], y_vals[yi + 1]
    z00, z10 = z[yi, xi], z[yi, xi + 1]
    z01, z11 = z[yi + 1, xi], z[yi + 1, xi + 1]
    if x0 == x1 or y0 == y1:
        return None
    t = (x - x0) / (x1 - x0)
    u = (y - y0) / (y1 - y0)
    return (1 - t) * (1 - u) * z00 + t * (1 - u) * z10 + (1 - t) * u * z01 + t * u * z11
