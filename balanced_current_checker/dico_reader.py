from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
from openpyxl import load_workbook

try:
    from can_equivalence import EquivalenceDb
except ImportError:
    EquivalenceDb = None


@dataclass
class ChannelDef:
    name: str
    units: str
    description: str
    min_val: float
    max_val: float
    raster: float
    resolution: float
    machine: str  # 'ME' or 'HSG'


@dataclass
class MachineGroup:
    machine: str          # 'ME' or 'HSG'
    channels: List[str]   # ordered [phase1, phase2, phase3]
    units: str
    min_val: float
    max_val: float
    raster: float
    resolution: float


@dataclass
class TorqueChannels:
    machine: str
    tq_cmd: str       # torque setpoint channel name
    tq_est: str       # torque estimated channel name
    speed: str        # rotor speed channel name
    voltage: str      # DC bus voltage channel name
    rotor_temp: str   # rotor temperature channel name
    sync_mode: str = ''  # channel name for 0=torque_mode / 1=speed_mode
    tq_cons: str = ''    # consolidated torque setpoint channel name
    norm_tq_der: str = ''  # normalized torque after derating
    norm_tq_sp: str = ''   # normalized torque setpoint (before derating)
    norm_tq_min: str = ''  # min normalized torque limit (aggregate)
    norm_tq_max: str = ''  # max normalized torque limit (aggregate)
    # Max-limit specific derating variables
    norm_tq_max_temp: str = ''
    norm_tq_max_udc: str = ''
    norm_tq_max_stop: str = ''
    norm_tq_cur_sat_max: str = ''
    norm_tq_max_asc_oc: str = ''
    norm_tq_max_cell_ov_fac: str = ''
    # Min-limit specific derating variables
    norm_tq_min_udc: str = ''
    norm_tq_min_back: str = ''
    norm_tq_min_cell_ov_fac: str = ''
    # Derating activation flag channels (Vbx_*)
    spt_der_act: str = ''
    spt_der_act_tqc: str = ''
    invt_stop: str = ''
    spt_der_v_oor: str = ''
    spt_der_ovh: str = ''
    max_cur_sat_act: str = ''
    asc_oc_der_act: str = ''
    invt_ztq_reg_act: str = ''
    dm_ov_sfty: str = ''
    back_der_act: str = ''
    # EAJS (Curative Antijerk) activation flag
    eajs_cor_act: str = ''
    # Saturation voltage channels
    v_d: str = ''
    v_q: str = ''
    v_d_ang: str = ''
    v_d_harm: str = ''
    v_q_ang: str = ''
    v_q_harm: str = ''
    rgl_sat: str = ''
    # Current channels (Id, Iq)
    i_d: str = ''
    i_q: str = ''
    i_d_req: str = ''
    i_q_req: str = ''


# DICO variable descriptions (loaded once from Excel)
DICO_DESCRIPTIONS: Dict[str, str] = {}


# Known variable name patterns for torque analysis
# (DICO names are used to validate existence; actual MF4 channel names used for loading)
_TQ_CMD_ME = 'Vxx_emot_tq_sp'
_TQ_CMD_HSG = 'Vxx_emot_tq_sp_emot2'
_TQ_EST_ME = 'Wxx_esti_emot_tq'
_TQ_EST_HSG = 'Wxx_esti_emot_tq_emot2'
_SPEED_ME = 'Wxx_emot_n'
_SPEED_HSG = 'Wxx_emot_n_emot2'
# DICO name user added -> actual MF4 name
_VOLTAGE_DICO = 'Wxx_udc_cs'
_VOLTAGE_MF4 = 'Wxx_udc_cs_ftr'
_TEMP_ME_DICO = 'Vxx_esti_rot_temp'
_TEMP_ME_MF4 = 'Vxx_esti_rot_temp_raw'
_TEMP_HSG_DICO = 'Vxx_esti_rot_temp_emot2'
_TEMP_HSG_MF4 = 'Vxx_esti_rot_temp_raw_emot2'
_SYNC_MODE_ME = 'Wbx_spt_sync_act_cs'
_SYNC_MODE_HSG = 'Wbx_spt_sync_act_cs_emot2'
_TQ_CONS_ME = 'Vxx_cs_tq_sp'
_TQ_CONS_HSG = 'Vxx_cs_tq_sp_emot2'
_NORM_TQ_DER_ME = 'Vxx_norm_tq_sp_der'
_NORM_TQ_DER_HSG = 'Vxx_norm_tq_sp_der_emot2'
_NORM_TQ_SP_ME = 'Vxx_norm_tq_sp'
_NORM_TQ_SP_HSG = 'Vxx_norm_tq_sp_emot2'
_NORM_TQ_MIN_ME = 'Vxx_norm_tq_min'
_NORM_TQ_MIN_HSG = 'Vxx_norm_tq_min_emot2'
_NORM_TQ_MAX_ME = 'Vxx_norm_tq_max'
_NORM_TQ_MAX_HSG = 'Vxx_norm_tq_max_emot2'
# Max-limit specific derating
_NORM_TQ_MAX_TEMP_ME = 'Vxx_norm_tq_max_temp'
_NORM_TQ_MAX_TEMP_HSG = 'Vxx_norm_tq_max_temp_emot2'
_NORM_TQ_MAX_UDC_ME = 'Vxx_norm_tq_max_udc'
_NORM_TQ_MAX_UDC_HSG = 'Vxx_norm_tq_max_udc_emot2'
_NORM_TQ_MAX_STOP_ME = 'Vxx_norm_tq_max_stop'
_NORM_TQ_MAX_STOP_HSG = 'Vxx_norm_tq_max_stop_emot2'
_NORM_TQ_CUR_SAT_MAX_ME = 'Vxx_norm_tq_cur_sat_max'
_NORM_TQ_CUR_SAT_MAX_HSG = 'Vxx_norm_tq_cur_sat_max_emot2'
_NORM_TQ_MAX_ASC_OC_ME = 'Vxx_norm_tq_max_asc_oc'
_NORM_TQ_MAX_ASC_OC_HSG = 'Vxx_norm_tq_max_asc_oc_emot2'
_NORM_TQ_MAX_CELL_OV_FAC_ME = 'Vxx_norm_tq_max_cell_ov_fac'
_NORM_TQ_MAX_CELL_OV_FAC_HSG = 'Vxx_norm_tq_max_cell_ov_fac_emot2'
# Min-limit specific derating
_NORM_TQ_MIN_UDC_ME = 'Vxx_norm_tq_min_udc'
_NORM_TQ_MIN_UDC_HSG = 'Vxx_norm_tq_min_udc_emot2'
_NORM_TQ_MIN_BACK_ME = 'Vxx_norm_tq_min_back'
_NORM_TQ_MIN_BACK_HSG = 'Vxx_norm_tq_min_back_emot2'
_NORM_TQ_MIN_CELL_OV_FAC_ME = 'Vxx_norm_tq_min_cell_ov_fac'
_NORM_TQ_MIN_CELL_OV_FAC_HSG = 'Vxx_norm_tq_min_cell_ov_fac_emot2'
# Derating activation flags (Vbx_*)
_SPT_DER_ACT_ME = 'Vbx_spt_der_act'
_SPT_DER_ACT_HSG = 'Vbx_spt_der_act_emot2'
_SPT_DER_ACT_TQC_ME = 'Vbx_spt_der_act_tqc'
_SPT_DER_ACT_TQC_HSG = 'Vbx_spt_der_act_tqc_emot2'
_INVT_STOP_ME = 'Vbx_invt_stop'
_INVT_STOP_HSG = 'Vbx_invt_stop_emot2'
_SPT_DER_V_OOR_ME = 'Vbx_spt_der_v_oor'
_SPT_DER_V_OOR_HSG = 'Vbx_spt_der_v_oor_emot2'
_SPT_DER_OVH_ME = 'Vbx_spt_der_ovh'
_SPT_DER_OVH_HSG = 'Vbx_spt_der_ovh_emot2'
_MAX_CUR_SAT_ACT_ME = 'Vbx_max_cur_sat_act'
_MAX_CUR_SAT_ACT_HSG = 'Vbx_max_cur_sat_act_emot2'
_ASC_OC_DER_ACT_ME = 'Vbx_asc_oc_der_act'
_ASC_OC_DER_ACT_HSG = 'Vbx_asc_oc_der_act_emot2'
_INVT_ZTQ_REG_ACT_ME = 'Vbx_invt_ztq_reg_act'
_INVT_ZTQ_REG_ACT_HSG = 'Vbx_invt_ztq_reg_act_emot2'
_DM_OV_SFTY_ME = 'Vbx_dm_ov_sfty'
_DM_OV_SFTY_HSG = 'Vbx_dm_ov_sfty_emot2'
_BACK_DER_ACT_ME = 'Vbx_back_der_act'
_BACK_DER_ACT_HSG = 'Vbx_back_der_act_emot2'
# EAJS (Curative Antijerk) activation
_EAJS_COR_ACT_ME = 'Vbx_hevc_eajs_cor_act'
_EAJS_COR_ACT_HSG = 'Vbx_hevc_eajs_cor_act_emot2'
# Saturation voltage channels
_V_D_ME = 'Wxx_v_d'
_V_D_HSG = 'Wxx_v_d_emot2'
_V_Q_ME = 'Wxx_v_q'
_V_Q_HSG = 'Wxx_v_q_emot2'
_V_D_ANG_ME = 'Wxx_v_d_ang'
_V_D_ANG_HSG = 'Wxx_v_d_ang_emot2'
_V_D_HARM_ME = 'Wxx_v_d_harm'
_V_D_HARM_HSG = 'Wxx_v_d_harm_emot2'
_V_Q_ANG_ME = 'Wxx_v_q_ang'
_V_Q_ANG_HSG = 'Wxx_v_q_ang_emot2'
_V_Q_HARM_ME = 'Wxx_v_q_harm'
_V_Q_HARM_HSG = 'Wxx_v_q_harm_emot2'
_RGL_SAT_ME = 'Wbx_rgl_sat'
_RGL_SAT_HSG = 'Wbx_rgl_sat_emot2'
# Current channels (Id, Iq measured and required)
_I_D_ME = 'Wxx_i_d'
_I_D_HSG = 'Wxx_i_d_emot2'
_I_Q_ME = 'Wxx_i_q'
_I_Q_HSG = 'Wxx_i_q_emot2'
_I_D_REQ_ME = 'Wxx_i_d_req'
_I_D_REQ_HSG = 'Wxx_i_d_req_emot2'
_I_Q_REQ_ME = 'Wxx_i_q_req'
_I_Q_REQ_HSG = 'Wxx_i_q_req_emot2'
# Power balance variables (ME inverter)
_EMOT1_POW_ME = 'Vxx_emot1_pow'
_INV1_I_ME = 'Vxx_inv1_i_mes_mux'
_HVBUS_V_INV1_ME = 'Vxx_hvbus_v_inv1_mux'
_EST_TQ_EMOT1_ME = 'Vxx_est_emot1_tq'
_SPD_EMOT1_ME = 'Vxx_emot1_spd_rpm'
# Power balance variables (HSG inverter)
_EMOT2_POW_HSG = 'Vxx_emot2_pow'
_INV2_I_HSG = 'Vxx_inv2_i_mes_mux'
_HVBUS_V_INV2_HSG = 'Vxx_hvbus_v_inv2_mux'
_EST_TQ_EMOT2_HSG = 'Vxx_est_emot2_tq'
_SPD_EMOT2_HSG = 'Vxx_emot2_spd_rpm'
# Global power balance variables
_HVB_POW = 'Vxx_hvb_pow'
_HVBUS_POW_CONS_EST = 'Vxx_hvbus_pow_cons_est'
_HVBUS_AFTR_RLY_V = 'Vxx_hvbus_aftr_rly_v_100ms'
# ICE power variables
_ICE_TQ = 'Vxx_eng_crksft_intm_arb_raw_tq'
_ICE_SPD = 'Vxx_avg_eng_spd'
# DLS (Drive Line Status) variables
_VNX_CRT_VH_DL = 'Vnx_crt_vh_dl'
_VNX_TG_VH_DL = 'Vnx_tg_vh_dl'
# Wheel speed and radius
_WHL_SPD = 'Vxx_whl_vh_spd'
_WHEEL_RADIUS_M = 0.3432

_DLS_TABLE: Dict[int, tuple] = {}  # dls_code -> (name, ratio_ice, ratio_me)


def _dico_to_mf4(name: str, name_mf4_map: dict) -> str:
    return name_mf4_map.get(name, name)


def _torque_channel(
    defined: set, dico_name: str, mf4_name: str, dico_to_mf4_map: dict, fallback: str = ''
) -> str:
    if dico_name in defined:
        return dico_to_mf4_map.get(dico_name, dico_name)
    return mf4_name if mf4_name and mf4_name in defined else fallback


def read_torque_channels(filepath: str) -> Dict[str, TorqueChannels]:
    wb = load_workbook(filepath, data_only=True)
    ws_name = 'DICO_Variables' if 'DICO_Variables' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[ws_name]
    defined = set()
    descriptions: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if name is not None and isinstance(name, str):
            defined.add(name)
            if len(row) > 2 and row[2] is not None:
                descriptions[name] = str(row[2]).strip()
    DICO_DESCRIPTIONS.update(descriptions)

    # Map DICO names to actual MF4 names
    dico_to_mf4_map = {
        _VOLTAGE_DICO: _VOLTAGE_MF4,
        _TEMP_ME_DICO: _TEMP_ME_MF4,
        _TEMP_HSG_DICO: _TEMP_HSG_MF4,
    }

    def _resolve(n: str) -> str:
        return dico_to_mf4_map.get(n, n)

    result = {}
    if _TQ_CMD_ME in defined and _SPEED_ME in defined and _TEMP_ME_DICO in defined:
        result['ME'] = TorqueChannels(
            machine='ME', tq_cmd=_TQ_CMD_ME,
            tq_est=_TQ_EST_ME if _TQ_EST_ME in defined else '',
            speed=_SPEED_ME,
            voltage=_resolve(_VOLTAGE_DICO) if _VOLTAGE_DICO in defined else '',
            rotor_temp=_resolve(_TEMP_ME_DICO),
            sync_mode=_SYNC_MODE_ME if _SYNC_MODE_ME in defined else '',
            tq_cons=_TQ_CONS_ME if _TQ_CONS_ME in defined else '',
            norm_tq_der=_NORM_TQ_DER_ME if _NORM_TQ_DER_ME in defined else '',
            norm_tq_sp=_NORM_TQ_SP_ME if _NORM_TQ_SP_ME in defined else '',
            norm_tq_min=_NORM_TQ_MIN_ME if _NORM_TQ_MIN_ME in defined else '',
            norm_tq_max=_NORM_TQ_MAX_ME if _NORM_TQ_MAX_ME in defined else '',
            norm_tq_max_temp=_NORM_TQ_MAX_TEMP_ME if _NORM_TQ_MAX_TEMP_ME in defined else '',
            norm_tq_max_udc=_NORM_TQ_MAX_UDC_ME if _NORM_TQ_MAX_UDC_ME in defined else '',
            norm_tq_max_stop=_NORM_TQ_MAX_STOP_ME if _NORM_TQ_MAX_STOP_ME in defined else '',
            norm_tq_cur_sat_max=_NORM_TQ_CUR_SAT_MAX_ME if _NORM_TQ_CUR_SAT_MAX_ME in defined else '',
            norm_tq_max_asc_oc=_NORM_TQ_MAX_ASC_OC_ME if _NORM_TQ_MAX_ASC_OC_ME in defined else '',
            norm_tq_max_cell_ov_fac=_NORM_TQ_MAX_CELL_OV_FAC_ME if _NORM_TQ_MAX_CELL_OV_FAC_ME in defined else '',
            norm_tq_min_udc=_NORM_TQ_MIN_UDC_ME if _NORM_TQ_MIN_UDC_ME in defined else '',
            norm_tq_min_back=_NORM_TQ_MIN_BACK_ME if _NORM_TQ_MIN_BACK_ME in defined else '',
            norm_tq_min_cell_ov_fac=_NORM_TQ_MIN_CELL_OV_FAC_ME if _NORM_TQ_MIN_CELL_OV_FAC_ME in defined else '',
            spt_der_act=_SPT_DER_ACT_ME if _SPT_DER_ACT_ME in defined else '',
            spt_der_act_tqc=_SPT_DER_ACT_TQC_ME if _SPT_DER_ACT_TQC_ME in defined else '',
            invt_stop=_INVT_STOP_ME if _INVT_STOP_ME in defined else '',
            spt_der_v_oor=_SPT_DER_V_OOR_ME if _SPT_DER_V_OOR_ME in defined else '',
            spt_der_ovh=_SPT_DER_OVH_ME if _SPT_DER_OVH_ME in defined else '',
            max_cur_sat_act=_MAX_CUR_SAT_ACT_ME if _MAX_CUR_SAT_ACT_ME in defined else '',
            asc_oc_der_act=_ASC_OC_DER_ACT_ME if _ASC_OC_DER_ACT_ME in defined else '',
            invt_ztq_reg_act=_INVT_ZTQ_REG_ACT_ME if _INVT_ZTQ_REG_ACT_ME in defined else '',
            dm_ov_sfty=_DM_OV_SFTY_ME if _DM_OV_SFTY_ME in defined else '',
            back_der_act=_BACK_DER_ACT_ME if _BACK_DER_ACT_ME in defined else '',
            eajs_cor_act=_EAJS_COR_ACT_ME if _EAJS_COR_ACT_ME in defined else '',
            v_d=_V_D_ME if _V_D_ME in defined else '',
            v_q=_V_Q_ME if _V_Q_ME in defined else '',
            v_d_ang=_V_D_ANG_ME if _V_D_ANG_ME in defined else '',
            v_d_harm=_V_D_HARM_ME if _V_D_HARM_ME in defined else '',
            v_q_ang=_V_Q_ANG_ME if _V_Q_ANG_ME in defined else '',
            v_q_harm=_V_Q_HARM_ME if _V_Q_HARM_ME in defined else '',
            rgl_sat=_RGL_SAT_ME if _RGL_SAT_ME in defined else '',
            i_d=_I_D_ME if _I_D_ME in defined else '',
            i_q=_I_Q_ME if _I_Q_ME in defined else '',
            i_d_req=_I_D_REQ_ME if _I_D_REQ_ME in defined else '',
            i_q_req=_I_Q_REQ_ME if _I_Q_REQ_ME in defined else '',
        )
    if _TQ_CMD_HSG in defined and _SPEED_HSG in defined and _TEMP_HSG_DICO in defined:
        result['HSG'] = TorqueChannels(
            machine='HSG', tq_cmd=_TQ_CMD_HSG,
            tq_est=_TQ_EST_HSG if _TQ_EST_HSG in defined else '',
            speed=_SPEED_HSG,
            voltage=_resolve(_VOLTAGE_DICO) if _VOLTAGE_DICO in defined else '',
            rotor_temp=_resolve(_TEMP_HSG_DICO),
            sync_mode=_SYNC_MODE_HSG if _SYNC_MODE_HSG in defined else '',
            tq_cons=_TQ_CONS_HSG if _TQ_CONS_HSG in defined else '',
            norm_tq_der=_NORM_TQ_DER_HSG if _NORM_TQ_DER_HSG in defined else '',
            norm_tq_sp=_NORM_TQ_SP_HSG if _NORM_TQ_SP_HSG in defined else '',
            norm_tq_min=_NORM_TQ_MIN_HSG if _NORM_TQ_MIN_HSG in defined else '',
            norm_tq_max=_NORM_TQ_MAX_HSG if _NORM_TQ_MAX_HSG in defined else '',
            norm_tq_max_temp=_NORM_TQ_MAX_TEMP_HSG if _NORM_TQ_MAX_TEMP_HSG in defined else '',
            norm_tq_max_udc=_NORM_TQ_MAX_UDC_HSG if _NORM_TQ_MAX_UDC_HSG in defined else '',
            norm_tq_max_stop=_NORM_TQ_MAX_STOP_HSG if _NORM_TQ_MAX_STOP_HSG in defined else '',
            norm_tq_cur_sat_max=_NORM_TQ_CUR_SAT_MAX_HSG if _NORM_TQ_CUR_SAT_MAX_HSG in defined else '',
            norm_tq_max_asc_oc=_NORM_TQ_MAX_ASC_OC_HSG if _NORM_TQ_MAX_ASC_OC_HSG in defined else '',
            norm_tq_max_cell_ov_fac=_NORM_TQ_MAX_CELL_OV_FAC_HSG if _NORM_TQ_MAX_CELL_OV_FAC_HSG in defined else '',
            norm_tq_min_udc=_NORM_TQ_MIN_UDC_HSG if _NORM_TQ_MIN_UDC_HSG in defined else '',
            norm_tq_min_back=_NORM_TQ_MIN_BACK_HSG if _NORM_TQ_MIN_BACK_HSG in defined else '',
            norm_tq_min_cell_ov_fac=_NORM_TQ_MIN_CELL_OV_FAC_HSG if _NORM_TQ_MIN_CELL_OV_FAC_HSG in defined else '',
            spt_der_act=_SPT_DER_ACT_HSG if _SPT_DER_ACT_HSG in defined else '',
            spt_der_act_tqc=_SPT_DER_ACT_TQC_HSG if _SPT_DER_ACT_TQC_HSG in defined else '',
            invt_stop=_INVT_STOP_HSG if _INVT_STOP_HSG in defined else '',
            spt_der_v_oor=_SPT_DER_V_OOR_HSG if _SPT_DER_V_OOR_HSG in defined else '',
            spt_der_ovh=_SPT_DER_OVH_HSG if _SPT_DER_OVH_HSG in defined else '',
            max_cur_sat_act=_MAX_CUR_SAT_ACT_HSG if _MAX_CUR_SAT_ACT_HSG in defined else '',
            asc_oc_der_act=_ASC_OC_DER_ACT_HSG if _ASC_OC_DER_ACT_HSG in defined else '',
            invt_ztq_reg_act=_INVT_ZTQ_REG_ACT_HSG if _INVT_ZTQ_REG_ACT_HSG in defined else '',
            dm_ov_sfty=_DM_OV_SFTY_HSG if _DM_OV_SFTY_HSG in defined else '',
            back_der_act=_BACK_DER_ACT_HSG if _BACK_DER_ACT_HSG in defined else '',
            eajs_cor_act=_EAJS_COR_ACT_HSG if _EAJS_COR_ACT_HSG in defined else '',
            v_d=_V_D_HSG if _V_D_HSG in defined else '',
            v_q=_V_Q_HSG if _V_Q_HSG in defined else '',
            v_d_ang=_V_D_ANG_HSG if _V_D_ANG_HSG in defined else '',
            v_d_harm=_V_D_HARM_HSG if _V_D_HARM_HSG in defined else '',
            v_q_ang=_V_Q_ANG_HSG if _V_Q_ANG_HSG in defined else '',
            v_q_harm=_V_Q_HARM_HSG if _V_Q_HARM_HSG in defined else '',
            rgl_sat=_RGL_SAT_HSG if _RGL_SAT_HSG in defined else '',
            i_d=_I_D_HSG if _I_D_HSG in defined else '',
            i_q=_I_Q_HSG if _I_Q_HSG in defined else '',
            i_d_req=_I_D_REQ_HSG if _I_D_REQ_HSG in defined else '',
            i_q_req=_I_Q_REQ_HSG if _I_Q_REQ_HSG in defined else '',
        )
    return result


def read_dico(filepath: str) -> Dict[str, MachineGroup]:
    wb = load_workbook(filepath, data_only=True)
    ws_name = 'DICO_Variables' if 'DICO_Variables' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[ws_name]

    # Map variable name -> raw row data
    raw_channels: List[ChannelDef] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if name is None or not isinstance(name, str):
            continue
        # Only actual phase current variables (Wxx_i_mot_ph_*), not safety (Wsx_*)
        if not name.startswith('Wxx') or 'mot_ph' not in name:
            continue
        raw_channels.append(ChannelDef(
            name=name,
            units=str(row[1] or ''),
            description=str(row[2] or ''),
            min_val=float(row[3]) if row[3] is not None else -500.0,
            max_val=float(row[4]) if row[4] is not None else 500.0,
            raster=float(row[5]) if row[5] is not None else 0.001,
            resolution=float(row[6]) if row[6] is not None else 0.015625,
            machine='HSG' if '_emot2' in name else 'ME',
        ))

    result: Dict[str, MachineGroup] = {}
    for machine in ('ME', 'HSG'):
        chs = [c for c in raw_channels if c.machine == machine]
        if machine == 'ME':
            phases = sorted([c for c in chs if not c.name.endswith('_emot2')],
                            key=lambda x: x.name)
        else:
            phases = sorted([c for c in chs if c.name.endswith('_emot2')],
                            key=lambda x: x.name)
        # Expect exactly 3 phases per machine
        if not phases:
            continue
        result[machine] = MachineGroup(
            machine=machine,
            channels=[p.name for p in phases],
            units=phases[0].units,
            min_val=phases[0].min_val,
            max_val=phases[0].max_val,
            raster=phases[0].raster,
            resolution=phases[0].resolution,
        )

    return result


def read_equivalences_from_dico(filepath: str) -> Optional['EquivalenceDb']:
    if EquivalenceDb is None:
        return None
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception:
        return None

    if 'Equivalencias_CAN' not in wb.sheetnames:
        return None
    ws = wb['Equivalencias_CAN']

    db = EquivalenceDb()
    skip_vals = {'-', 'Not Consumed', 'Not Transmitted', ''}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        basic_me = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
        basic_hsg = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
        var_int = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''

        if not var_int or var_int in skip_vals:
            continue
        if basic_me and basic_me not in skip_vals and basic_me != 'HSG only':
            db.add(basic_me, var_int)
        if basic_hsg and basic_hsg not in skip_vals and basic_hsg != 'ME only':
            db.add(basic_hsg, var_int)

    return db if len(db) > 0 else None


def read_dls_table(filepath: str) -> Dict[int, tuple]:
    """Read Marchas_DLS sheet. Returns {dls_code: (name, ratio_ice, ratio_me)}."""
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception:
        return {}
    if 'Marchas_DLS' not in wb.sheetnames:
        return {}
    ws = wb['Marchas_DLS']
    table: Dict[int, tuple] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 10:
            continue
        name = row[1]
        dls_code = row[9]  # col J = DLS final
        ratio_ice = row[7]  # col H
        ratio_me = row[8]   # col I
        if dls_code is None or name is None:
            continue
        try:
            code = int(dls_code)
            ri = float(ratio_ice) if ratio_ice is not None else 0.0
            rm = float(ratio_me) if ratio_me is not None else 0.0
        except (ValueError, TypeError):
            continue
        table[code] = (str(name).strip(), ri, rm)
    _DLS_TABLE.update(table)
    return table
