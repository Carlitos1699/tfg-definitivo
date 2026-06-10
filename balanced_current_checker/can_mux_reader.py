from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass


@dataclass
class CanSignal:
    signal_name: str
    basic_name: str
    unit: str
    resolution: float
    offset: float
    min_val: float
    max_val: float
    period_ms: int
    machine: str
    direction: str


ARCHITECTURES = ['Sweet200', 'Sweet400']


def _read_sweet400(path: str) -> Dict[str, List[CanSignal]]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['INV_PEBK']

    result: Dict[str, List[CanSignal]] = {'ME': [], 'HSG': []}

    for row in ws.iter_rows(min_row=2, values_only=True):
        signal_name = row[0]
        if signal_name is None:
            continue
        bme = row[13] if len(row) > 13 else ''
        basic_name_me = str(bme).strip() if bme else ''
        bhsg = row[14] if len(row) > 14 else ''
        basic_name_hsg = str(bhsg).strip() if bhsg else ''
        unit = str(row[5] or '') if len(row) > 5 else ''
        raw_res = row[6] if len(row) > 6 else None
        resolution = float(raw_res) if raw_res is not None and str(raw_res).strip() != '' else 1.0
        raw_off = row[7] if len(row) > 7 else None
        offset = float(raw_off) if raw_off is not None and str(raw_off).strip() != '' else 0.0
        raw_min = row[8] if len(row) > 8 else None
        min_val = float(raw_min) if raw_min is not None and str(raw_min).strip() != '' else 0.0
        raw_max = row[9] if len(row) > 9 else None
        max_val = float(raw_max) if raw_max is not None and str(raw_max).strip() != '' else 0.0
        r4 = row[4] if len(row) > 4 else None
        period = int(float(str(r4))) if r4 is not None and str(r4).strip() not in ('', '-', 'None') else 0
        dme = row[19] if len(row) > 19 else None
        dir_me = str(dme).strip() if dme else ''
        dhsg = row[20] if len(row) > 20 else None
        dir_hsg = str(dhsg).strip() if dhsg else ''

        exclude_vals = {'-', 'ME only', 'HSG only', 'Not Consumed', ''}
        if basic_name_me not in exclude_vals:
            result['ME'].append(CanSignal(
                signal_name=str(signal_name), basic_name=basic_name_me,
                unit=unit, resolution=resolution, offset=offset,
                min_val=min_val, max_val=max_val, period_ms=period,
                machine='ME', direction=dir_me))
        if basic_name_hsg not in exclude_vals:
            result['HSG'].append(CanSignal(
                signal_name=str(signal_name), basic_name=basic_name_hsg,
                unit=unit, resolution=resolution, offset=offset,
                min_val=min_val, max_val=max_val, period_ms=period,
                machine='HSG', direction=dir_hsg))

    return result


def _read_sweet200_me(path: str) -> List[CanSignal]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['INV_ME']

    signals = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        signal_name = row[0]
        if signal_name is None:
            continue
        bn = row[14] if len(row) > 14 else ''
        basic_name = str(bn).strip() if bn else ''
        if not basic_name or basic_name in ('-', 'Not Consumed', 'ME only', 'HSG only'):
            continue
        unit = str(row[6] or '') if len(row) > 6 else ''
        r7 = row[7] if len(row) > 7 else None
        resolution = float(r7) if r7 is not None and str(r7).strip() != '' else 1.0
        r8 = row[8] if len(row) > 8 else None
        offset = float(r8) if r8 is not None and str(r8).strip() != '' else 0.0
        r9 = row[9] if len(row) > 9 else None
        min_val = float(r9) if r9 is not None and str(r9).strip() != '' else 0.0
        r10 = row[10] if len(row) > 10 else None
        max_val = float(r10) if r10 is not None and str(r10).strip() != '' else 0.0
        r3 = row[3] if len(row) > 3 else None
        period = int(float(str(r3))) if r3 is not None and str(r3).strip() not in ('', '-', 'None') else 0
        direction = str(row[17]).strip() if len(row) > 17 and row[17] is not None else ''

        signals.append(CanSignal(
            signal_name=str(signal_name), basic_name=str(basic_name),
            unit=unit, resolution=resolution, offset=offset,
            min_val=min_val, max_val=max_val, period_ms=period,
            machine='ME', direction=direction))
    return signals


def _read_sweet200_hsg(path: str) -> List[CanSignal]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['INV_HSG']

    signals = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        signal_name = row[0]
        if signal_name is None:
            continue
        bn = row[14] if len(row) > 14 else ''
        basic_name = str(bn).strip() if bn else ''
        if not basic_name or basic_name in ('-', 'Not Consumed', 'ME only', 'HSG only'):
            continue
        unit = str(row[6] or '') if len(row) > 6 else ''
        r7 = row[7] if len(row) > 7 else None
        resolution = float(r7) if r7 is not None and str(r7).strip() != '' else 1.0
        r8 = row[8] if len(row) > 8 else None
        offset = float(r8) if r8 is not None and str(r8).strip() != '' else 0.0
        r9 = row[9] if len(row) > 9 else None
        min_val = float(r9) if r9 is not None and str(r9).strip() != '' else 0.0
        r10 = row[10] if len(row) > 10 else None
        max_val = float(r10) if r10 is not None and str(r10).strip() != '' else 0.0
        r3 = row[3] if len(row) > 3 else None
        period = int(float(str(r3))) if r3 is not None and str(r3).strip() not in ('', '-', 'None') else 0
        direction = str(row[17]).strip() if len(row) > 17 and row[17] is not None else ''

        signals.append(CanSignal(
            signal_name=str(signal_name), basic_name=str(basic_name),
            unit=unit, resolution=resolution, offset=offset,
            min_val=min_val, max_val=max_val, period_ms=period,
            machine='HSG', direction=direction))
    return signals


def read_can_mux(architecture: str, paths: dict) -> Dict[str, List[CanSignal]]:
    if architecture == 'Sweet400':
        return _read_sweet400(paths['mux'])
    elif architecture == 'Sweet200':
        me_sigs = _read_sweet200_me(paths['me'])
        hsg_sigs = _read_sweet200_hsg(paths['hsg'])
        return {'ME': me_sigs, 'HSG': hsg_sigs}
    else:
        return {'ME': [], 'HSG': []}
